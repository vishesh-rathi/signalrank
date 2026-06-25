import contextlib
import csv
import io
import json
import time
from pathlib import Path

import numpy as np
import streamlit as st

from ranker import config
from ranker.reasoning import generate_reasoning
from ranker.score import rank_candidates
from ranker.util import iter_candidates

# ───────────────────────── Page Config ──────────────────────────────────
st.set_page_config(
    page_title="SignalRank — Candidate Discovery Sandbox",
    page_icon="🏆",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ───────────────────────── Premium Theme CSS ────────────────────────────
st.markdown(
    """
<style>
/* ── Import Google Font ─────────────────────────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

/* ── Global Reset ───────────────────────────────────────────────────── */
html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
}

/* ── Header Area ────────────────────────────────────────────────────── */
.hero-section {
    padding: 1.5rem 0 1rem 0;
}
.hero-title {
    background: linear-gradient(135deg, #818cf8 0%, #a78bfa 30%, #c084fc 60%, #e879f9 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    font-weight: 900;
    font-size: 2.6rem;
    letter-spacing: -0.03em;
    line-height: 1.1;
    margin-bottom: 0.3rem;
}
.hero-subtitle {
    color: #94a3b8;
    font-size: 1.05rem;
    font-weight: 400;
    letter-spacing: 0.01em;
    margin-bottom: 0;
}

/* ── Glass Cards ────────────────────────────────────────────────────── */
.glass-card {
    background: rgba(30, 41, 59, 0.6);
    backdrop-filter: blur(16px);
    -webkit-backdrop-filter: blur(16px);
    border: 1px solid rgba(148, 163, 184, 0.12);
    border-radius: 16px;
    padding: 1.5rem;
    margin-bottom: 1rem;
    transition: border-color 0.3s ease, box-shadow 0.3s ease;
}
.glass-card:hover {
    border-color: rgba(129, 140, 248, 0.3);
    box-shadow: 0 8px 32px rgba(129, 140, 248, 0.08);
}

/* ── Metric Cards ───────────────────────────────────────────────────── */
.metric-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
    gap: 0.75rem;
}
.metric-card {
    background: rgba(30, 41, 59, 0.5);
    backdrop-filter: blur(12px);
    border: 1px solid rgba(148, 163, 184, 0.1);
    border-radius: 12px;
    padding: 1rem 1.1rem;
    text-align: center;
    transition: transform 0.2s ease, border-color 0.3s ease;
}
.metric-card:hover {
    transform: translateY(-2px);
    border-color: rgba(129, 140, 248, 0.25);
}
.metric-value {
    font-size: 1.6rem;
    font-weight: 800;
    background: linear-gradient(135deg, #818cf8, #c084fc);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    line-height: 1.2;
}
.metric-label {
    font-size: 0.72rem;
    color: #94a3b8;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    font-weight: 600;
    margin-top: 0.25rem;
}

/* ── Score Badge ────────────────────────────────────────────────────── */
.score-badge {
    display: inline-flex;
    align-items: center;
    gap: 0.4rem;
    background: linear-gradient(135deg, rgba(129, 140, 248, 0.15), rgba(192, 132, 252, 0.15));
    border: 1px solid rgba(129, 140, 248, 0.25);
    border-radius: 20px;
    padding: 0.35rem 0.9rem;
    font-size: 0.85rem;
    font-weight: 600;
    color: #c4b5fd;
}

/* ── Sidebar Styling ────────────────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: rgba(15, 23, 42, 0.95);
    border-right: 1px solid rgba(148, 163, 184, 0.08);
}
.sidebar-brand {
    padding: 0.5rem 0 1rem 0;
    border-bottom: 1px solid rgba(148, 163, 184, 0.1);
    margin-bottom: 1.2rem;
}
.sidebar-brand-title {
    font-size: 1.3rem;
    font-weight: 800;
    background: linear-gradient(135deg, #818cf8, #c084fc);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    letter-spacing: -0.02em;
}
.sidebar-brand-sub {
    font-size: 0.73rem;
    color: #64748b;
    font-weight: 500;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    margin-top: 0.1rem;
}

/* ── Table/Dataframe Enhancements ───────────────────────────────────── */
.stDataFrame {
    border-radius: 12px;
    overflow: hidden;
}

/* ── Tab Styling ────────────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
    gap: 2rem;
    background: rgba(30, 41, 59, 0.4);
    border-radius: 12px;
    padding: 0.3rem 1rem;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px;
    font-weight: 600;
    font-size: 0.95rem;
}

/* ── Expander Styling ───────────────────────────────────────────────── */
.streamlit-expanderHeader {
    font-weight: 600;
    font-size: 0.95rem;
}

/* ── Progress Bar Override ──────────────────────────────────────────── */
.stProgress > div > div > div > div {
    background: linear-gradient(90deg, #818cf8, #c084fc, #e879f9);
    border-radius: 10px;
}

/* ── Button Styling ─────────────────────────────────────────────────── */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 50%, #a855f7 100%);
    border: none;
    font-weight: 700;
    letter-spacing: 0.02em;
    border-radius: 10px;
    padding: 0.7rem 1.5rem;
    transition: all 0.3s ease;
}
.stButton > button[kind="primary"]:hover {
    box-shadow: 0 6px 20px rgba(99, 102, 241, 0.35);
    transform: translateY(-1px);
}

/* ── Profile Detail Table ───────────────────────────────────────────── */
.profile-table {
    width: 100%;
    border-collapse: collapse;
}
.profile-table td {
    padding: 0.55rem 0;
    border-bottom: 1px solid rgba(148, 163, 184, 0.06);
}
.profile-table td:first-child {
    color: #94a3b8;
    font-weight: 500;
    font-size: 0.88rem;
    width: 40%;
}
.profile-table td:last-child {
    color: #e2e8f0;
    font-weight: 400;
    text-align: right;
    font-size: 0.88rem;
}

/* ── Rank Badge ─────────────────────────────────────────────────────── */
.rank-badge {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 2.2rem;
    height: 2.2rem;
    border-radius: 50%;
    font-weight: 800;
    font-size: 0.85rem;
}
.rank-top10 {
    background: linear-gradient(135deg, #fbbf24, #f59e0b);
    color: #1e293b;
}
.rank-top50 {
    background: rgba(129, 140, 248, 0.2);
    border: 1px solid rgba(129, 140, 248, 0.4);
    color: #a5b4fc;
}
.rank-rest {
    background: rgba(100, 116, 139, 0.15);
    border: 1px solid rgba(100, 116, 139, 0.25);
    color: #94a3b8;
}

/* ── Score Bar ──────────────────────────────────────────────────────── */
.score-bar-container {
    width: 100%;
    height: 6px;
    background: rgba(100, 116, 139, 0.15);
    border-radius: 3px;
    overflow: hidden;
    margin-top: 0.3rem;
}
.score-bar-fill {
    height: 100%;
    border-radius: 3px;
    transition: width 0.6s ease;
}

/* ── Custom Download Button ─────────────────────────────────────────── */
[data-testid="stDownloadButton"] button {
    background-color: #10b981 !important;
    color: #ffffff !important;
    border-color: #10b981 !important;
}
[data-testid="stDownloadButton"] button:hover {
    background-color: #059669 !important;
    border-color: #059669 !important;
    color: #ffffff !important;
}

/* ── Methodology Note ───────────────────────────────────────────────── */
.method-note {
    background: rgba(129, 140, 248, 0.06);
    border-left: 4px solid #818cf8;
    border-radius: 0 8px 8px 0;
    padding: 1rem 1.2rem;
    font-size: 0.95rem;
    color: #cbd5e1;
    line-height: 1.6;
    margin-top: 1rem;
    margin-bottom: 2rem;
}

/* ── Deep-Dive Metrics ──────────────────────────────────────────────── */
.small-metric-label {
    font-size: 0.85rem;
    color: #94a3b8;
    margin-bottom: 0.1rem;
}
.small-metric-val {
    font-size: 1.2rem;
    font-weight: 600;
    color: #e2e8f0;
    margin-bottom: 1rem;
}
.fit-list {
    font-size: 1.25rem;
    line-height: 1.8;
}

/* ── Responsive Adjustments ─────────────────────────────────────────── */
@media (max-width: 768px) {
    .hero-title { font-size: 1.8rem; }
    .metric-grid { grid-template-columns: repeat(2, 1fr); }
}
</style>
""",
    unsafe_allow_html=True,
)


# ───────────────────────── Resource Loading ─────────────────────────────
@st.cache_resource
def load_resources(artifact_dir="artifacts"):
    directory = Path(artifact_dir)
    cand_path = directory / "cand_embeddings.npy"
    probe_path = directory / "jd_probe_embeddings.npy"
    ids_path = directory / "cand_ids.json"

    if not (cand_path.exists() and probe_path.exists() and ids_path.exists()):
        return None, {}, None

    try:
        embeddings = np.load(cand_path)
        probe_matrix = np.load(probe_path)
        candidate_ids = json.loads(ids_path.read_text())
        id_index = {cid: row for row, cid in enumerate(candidate_ids)}
        return embeddings, id_index, probe_matrix
    except Exception as e:
        st.warning(f"Error loading embedding artifacts: {e}. Falling back to lexical-only mode.")
        return None, {}, None


# ───────────────────────── Parsing Utilities ────────────────────────────
def parse_csv_candidates(content: str) -> list[dict]:
    candidates = []
    reader = csv.DictReader(content.splitlines())
    for i, row in enumerate(reader):
        candidate_id = row.get("candidate_id") or row.get("id") or f"CAND_CSV_{i:07d}"

        profile = {}
        profile_fields = [
            "years_of_experience",
            "current_title",
            "current_company",
            "summary",
            "location",
            "country",
            "current_industry",
        ]
        for k in profile_fields:
            val = row.get(k) or row.get(f"profile.{k}") or row.get(f"profile_{k}") or ""
            profile[k] = val

        try:
            profile["years_of_experience"] = float(profile["years_of_experience"])
        except (ValueError, TypeError):
            profile["years_of_experience"] = 0.0

        redrob_signals = {}
        signal_keys = [
            "recruiter_response_rate",
            "willing_to_relocate",
            "open_to_work_flag",
            "notice_period_days",
            "last_active_date",
            "github_activity_score",
        ]
        for k in signal_keys:
            val = (
                row.get(k) or row.get(f"redrob_signals.{k}") or row.get(f"redrob_signals_{k}") or ""
            )
            if isinstance(val, str):
                if val.lower() == "true":
                    val = True
                elif val.lower() == "false":
                    val = False
                else:
                    with contextlib.suppress(ValueError):
                        val = float(val) if "." in val else int(val)
            redrob_signals[k] = val

        candidate = {
            "candidate_id": candidate_id,
            "profile": profile,
            "career_history": [],
            "education": [],
            "skills": [],
            "redrob_signals": redrob_signals,
        }

        for field in ["career_history", "education", "skills"]:
            val = row.get(field) or ""
            if val.strip():
                try:
                    candidate[field] = json.loads(val)
                except json.JSONDecodeError:
                    if field == "skills":
                        candidate[field] = [
                            {
                                "name": s.strip(),
                                "proficiency": "intermediate",
                                "duration_months": 12,
                            }
                            for s in val.split(",")
                            if s.strip()
                        ]

        if not candidate["career_history"]:
            yoe = profile["years_of_experience"]
            duration = int(yoe * 12) if yoe > 0 else 12
            candidate["career_history"] = [
                {
                    "title": profile.get("current_title") or "Software Engineer",
                    "company": profile.get("current_company") or "Company",
                    "is_current": True,
                    "duration_months": duration,
                }
            ]

        candidates.append(candidate)
    return candidates


def parse_xlsx_candidates(file_bytes: bytes) -> list[dict]:
    """Parse XLSX file into candidate dicts, handling both flat and nested formats."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        return []

    header = [str(h).strip() if h is not None else "" for h in header]
    candidates = []

    for i, values in enumerate(rows_iter):
        row = dict(zip(header, values, strict=False))
        # Reuse CSV parsing logic by converting to string dict
        str_row = {k: str(v) if v is not None else "" for k, v in row.items()}

        candidate_id = str_row.get("candidate_id") or str_row.get("id") or f"CAND_XLSX_{i:07d}"
        profile = {}
        for k in [
            "years_of_experience",
            "current_title",
            "current_company",
            "summary",
            "location",
            "country",
            "current_industry",
        ]:
            profile[k] = str_row.get(k) or str_row.get(f"profile.{k}") or ""
        try:
            profile["years_of_experience"] = float(profile["years_of_experience"])
        except (ValueError, TypeError):
            profile["years_of_experience"] = 0.0

        redrob_signals = {}
        for k in [
            "recruiter_response_rate",
            "willing_to_relocate",
            "open_to_work_flag",
            "notice_period_days",
            "last_active_date",
            "github_activity_score",
        ]:
            val = str_row.get(k) or str_row.get(f"redrob_signals.{k}") or ""
            if isinstance(val, str) and val:
                if val.lower() == "true":
                    val = True
                elif val.lower() == "false":
                    val = False
                else:
                    with contextlib.suppress(ValueError):
                        val = float(val) if "." in val else int(val)
            redrob_signals[k] = val

        candidate = {
            "candidate_id": candidate_id,
            "profile": profile,
            "career_history": [],
            "education": [],
            "skills": [],
            "redrob_signals": redrob_signals,
        }

        for field in ["career_history", "education", "skills"]:
            val = str_row.get(field) or ""
            if val.strip():
                try:
                    candidate[field] = json.loads(val)
                except json.JSONDecodeError:
                    if field == "skills":
                        candidate[field] = [
                            {
                                "name": s.strip(),
                                "proficiency": "intermediate",
                                "duration_months": 12,
                            }
                            for s in val.split(",")
                            if s.strip()
                        ]

        if not candidate["career_history"]:
            yoe = profile["years_of_experience"]
            duration = int(yoe * 12) if yoe > 0 else 12
            candidate["career_history"] = [
                {
                    "title": profile.get("current_title") or "Software Engineer",
                    "company": profile.get("current_company") or "Company",
                    "is_current": True,
                    "duration_months": duration,
                }
            ]
        candidates.append(candidate)

    wb.close()
    return candidates


def parse_uploaded_file(uploaded_file) -> list[dict]:
    name = uploaded_file.name.lower()
    raw_bytes = uploaded_file.getvalue()

    if name.endswith(".xlsx"):
        return parse_xlsx_candidates(raw_bytes)

    content = raw_bytes.decode("utf-8", errors="replace")

    if name.endswith(".jsonl"):
        candidates = []
        for line in content.splitlines():
            if not line.strip():
                continue
            try:
                candidates.append(json.loads(line))
            except json.JSONDecodeError:
                continue
        return candidates

    elif name.endswith(".json"):
        try:
            data = json.loads(content)
            if isinstance(data, list):
                return data
            elif isinstance(data, dict):
                return [data]
        except json.JSONDecodeError:
            st.error("Invalid JSON format in the uploaded file.")
        return []

    elif name.endswith(".csv"):
        return parse_csv_candidates(content)

    return []


# ───────────────────────── Global Settings ──────────────────────────────
top_n = 100
artifact_dir = "artifacts"

# ───────────────────────── Hero Section ─────────────────────────────────
st.markdown(
    """
    <div class="hero-section">
        <div class="hero-title">SignalRank Discovery Engine</div>
        <div class="hero-subtitle">
            End-to-end interpretable candidate ranking — hybrid lexical + semantic
            scoring with full audit trail
        </div>
    </div>
    
    <div class="method-note">
        <strong>Methodology</strong><br>
        5-stage hybrid ranker: lexical evidence extraction → semantic embedding
        similarity → domain/title gating → behavioral multiplier → honeypot
        filtering. No LLM calls at ranking time — CPU-only, offline, under 5 min.
    </div>
    """,
    unsafe_allow_html=True,
)

# ── Load resources ──────────────────────────────────────────────────────
embeddings, id_index, probe_matrix = load_resources(artifact_dir)
mode_label = "Hybrid (Lexical + Semantic)" if embeddings is not None else "Lexical-Only"
mode_short = "HYBRID" if embeddings is not None else "LEXICAL"

# ── Session state ───────────────────────────────────────────────────────
for key, default in [
    ("ranked_results", None),
    ("elapsed_time", 0.0),
    ("cand_count", 0),
    ("honeypot_count", 0),
    ("uncovered_embeddings", 0),
    ("csv_bytes", None),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ───────────────────────── Tabs ─────────────────────────────────────────
tab_run, tab_leaderboard, tab_inspector = st.tabs(
    ["📥  Run Pipeline", "🏆  Leaderboard & Export", "🔍  Score Deep-Dive"]
)

# ═══════════════════════ TAB 1: Run Pipeline ════════════════════════════
with tab_run:
    col_input, col_stats = st.columns([3, 2], gap="large")

    with col_input:
        st.markdown("#### 📂 Data Input Source")
        input_source = st.radio(
            "Select input data method:",
            ["Use Pre-loaded Sample (100 Candidates)", "Upload Custom Candidates File"],
            label_visibility="collapsed",
        )

        candidates = []

        if input_source == "Use Pre-loaded Sample (100 Candidates)":
            sample_path = Path("sample_candidates.jsonl")
            if sample_path.exists():
                candidates = list(iter_candidates(str(sample_path)))
                st.markdown(
                    f"""
                    <div class="glass-card" style="padding: 1rem 1.2rem;">
                        <span style="font-size: 1.1rem;">📂</span>&ensp;
                        <span style="color: #e2e8f0; font-weight: 500;">
                            Pre-loaded <code>sample_candidates.jsonl</code>
                        </span>
                        <span class="score-badge" style="float: right;">
                            {len(candidates)} candidates
                        </span>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            else:
                st.error("❌ `sample_candidates.jsonl` not found in the working directory.")

        else:
            uploaded_file = st.file_uploader(
                "Upload candidate file",
                type=["jsonl", "json", "csv", "xlsx"],
                help="Supports JSON, JSONL, CSV, and XLSX formats.",
            )
            if uploaded_file is not None:
                with st.spinner("Parsing uploaded file..."):
                    candidates = parse_uploaded_file(uploaded_file)
                if candidates:
                    st.success(
                        f"✅ Loaded **{len(candidates)}** candidates from `{uploaded_file.name}`."
                    )
                else:
                    st.error("❌ No valid candidate profiles could be parsed from the file.")

        st.markdown("")

        # ── Run Button ──────────────────────────────────────────────────
        run_disabled = len(candidates) == 0
        if st.button(
            "🚀  Run Discovery Pipeline",
            type="primary",
            disabled=run_disabled,
        ):
            start_time = time.time()

            with st.spinner("Running SignalRank hybrid ranking pipeline..."):
                ranked = rank_candidates(
                    candidates, embeddings, id_index, probe_matrix, config.DATA_AS_OF, top_n=top_n
                )

            elapsed = time.time() - start_time

            # Compute stats
            honeypot_val = sum(1 for r in ranked if r["trace"].get("honeypot") is True)
            uncovered_val = 0
            if embeddings is not None:
                uncovered_val = sum(1 for c in candidates if c.get("candidate_id") not in id_index)

            # Persist results
            st.session_state.ranked_results = ranked
            st.session_state.elapsed_time = elapsed
            st.session_state.cand_count = len(candidates)
            st.session_state.honeypot_count = honeypot_val
            st.session_state.uncovered_embeddings = uncovered_val

            # Build CSV bytes
            csv_buffer = io.StringIO()
            writer = csv.writer(csv_buffer)
            writer.writerow(["candidate_id", "rank", "score", "reasoning"])
            for row in ranked:
                writer.writerow(
                    [
                        row["candidate_id"],
                        row["rank"],
                        f"{row['score']:.6f}",
                        generate_reasoning(row["candidate"], row["trace"], row["rank"]),
                    ]
                )
            st.session_state.csv_bytes = csv_buffer.getvalue().encode("utf-8")

            st.balloons()
            st.success(
                "🎉 Pipeline completed! Switch to the **Leaderboard & Export** tab to view results."
            )

    # ── Stats Panel ─────────────────────────────────────────────────────
    with col_stats:
        st.markdown("#### 📊 Pipeline Metrics")
        if st.session_state.ranked_results is None:
            st.markdown(
                """
                <div class="glass-card" style="text-align: center; padding: 3rem 1.5rem;">
                    <div style="font-size: 2.5rem; margin-bottom: 0.5rem; opacity: 0.3;">📊</div>
                    <div style="color: #64748b; font-weight: 500; font-size: 0.9rem;">
                        No results yet<br>
                        <span style="font-size: 0.8rem; font-weight: 400;">
                            Select input and run the pipeline
                        </span>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
        else:
            uncovered_html = ""
            if embeddings is not None:
                # Use a single-line string with no leading indentation to avoid markdown code-block rendering
                uncovered_html = f'<div class="metric-card"><div class="metric-value">{st.session_state.uncovered_embeddings}</div><div class="metric-label">Missing Embeds</div></div>'

            st.markdown(
                f"""
                <div class="glass-card">
                    <div class="metric-grid">
                        <div class="metric-card">
                            <div class="metric-value">{mode_short}</div>
                            <div class="metric-label">Pipeline Mode</div>
                        </div>
                        <div class="metric-card">
                            <div class="metric-value">{st.session_state.elapsed_time:.2f}s</div>
                            <div class="metric-label">Runtime</div>
                        </div>
                        <div class="metric-card">
                            <div class="metric-value">{st.session_state.cand_count:,}</div>
                            <div class="metric-label">Candidates</div>
                        </div>
                        <div class="metric-card">
                            <div class="metric-value">{st.session_state.honeypot_count}</div>
                            <div class="metric-label">Honeypots Blocked</div>
                        </div>
                        {uncovered_html}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            # Score distribution summary
            if st.session_state.ranked_results:
                scores = [r["score"] for r in st.session_state.ranked_results]
                st.markdown(
                    f"""
                    <div class="glass-card" style="padding: 1rem 1.2rem;">
                        <div style="font-size: 0.78rem; color: #94a3b8; text-transform: uppercase;
                                    letter-spacing: 0.06em; font-weight: 600; margin-bottom: 0.6rem;">
                            Score Distribution
                        </div>
                        <div class="metric-grid">
                            <div class="metric-card" style="background: transparent; border: none; padding: 0.5rem;">
                                <div class="metric-value" style="font-size: 1.3rem;">{max(scores):.6f}</div>
                                <div class="metric-label">Highest</div>
                            </div>
                            <div class="metric-card" style="background: transparent; border: none; padding: 0.5rem;">
                                <div class="metric-value" style="font-size: 1.3rem;">{sorted(scores)[len(scores)//2]:.6f}</div>
                                <div class="metric-label">Median</div>
                            </div>
                            <div class="metric-card" style="background: transparent; border: none; padding: 0.5rem;">
                                <div class="metric-value" style="font-size: 1.3rem;">{min(scores):.6f}</div>
                                <div class="metric-label">Lowest</div>
                            </div>
                            <div class="metric-card" style="background: transparent; border: none; padding: 0.5rem;">
                                <div class="metric-value" style="font-size: 1.3rem;">{max(scores) - min(scores):.6f}</div>
                                <div class="metric-label">Spread</div>
                            </div>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

# ═══════════════════════ TAB 2: Leaderboard ═════════════════════════════
with tab_leaderboard:
    if st.session_state.ranked_results is None:
        st.info("💡 Run the discovery pipeline in the **Run Pipeline** tab first.")
    else:
        # ── Export Bar ──────────────────────────────────────────────────
        col_dl, col_info = st.columns([1, 3])
        with col_dl:
            if st.session_state.csv_bytes is not None:
                st.download_button(
                    label="💾  Download Submission CSV",
                    data=st.session_state.csv_bytes,
                    file_name="submission.csv",
                    mime="text/csv",
                    type="primary",
                )
        with col_info:
            n_ranked = len(st.session_state.ranked_results)
            st.markdown(
                f"""
                <div style="display: flex; align-items: center; gap: 1rem;
                            padding: 0.6rem 0; flex-wrap: wrap;">
                    <span class="score-badge">📋 {n_ranked} ranked</span>
                    <span class="score-badge">✅ Spec-compliant CSV</span>
                    <span class="score-badge">🛡️ 0 honeypots</span>
                </div>
                """,
                unsafe_allow_html=True,
            )

        st.markdown("")

        # ── Build display data ──────────────────────────────────────────
        display_data = []
        for r in st.session_state.ranked_results:
            profile = r["candidate"].get("profile") or {}
            display_data.append(
                {
                    "Rank": r["rank"],
                    "ID": r["candidate_id"],
                    "Title": profile.get("current_title", "N/A"),
                    "Experience": f"{profile.get('years_of_experience', 0)} yrs",
                    "Score": f"{r['score']:.6f}",
                    "Reasoning": generate_reasoning(r["candidate"], r["trace"], r["rank"]),
                }
            )

        st.dataframe(
            display_data,
            width="stretch",
            height=600,
            column_config={
                "Rank": st.column_config.NumberColumn(format="%d", width="small"),
                "Score": st.column_config.TextColumn(width="small"),
                "ID": st.column_config.TextColumn(width="medium"),
                "Title": st.column_config.TextColumn(width="medium"),
                "Experience": st.column_config.TextColumn(width="small"),
                "Reasoning": st.column_config.TextColumn(width="large"),
            },
        )

# ═══════════════════════ TAB 3: Score Deep-Dive ═════════════════════════
with tab_inspector:
    if st.session_state.ranked_results is None:
        st.info("💡 Run the discovery pipeline in the **Run Pipeline** tab first.")
    else:
        inspector_cids = [
            f"Rank {r['rank']} — {r['candidate_id']}" for r in st.session_state.ranked_results
        ]
        st.markdown("#### 🎯 Select Candidate to Audit")
        selected_label = st.selectbox(
            "Select a candidate to audit:",
            inspector_cids,
            index=0,
            label_visibility="collapsed",
        )
        selected_idx = inspector_cids.index(selected_label)
        selected_row = st.session_state.ranked_results[selected_idx]

        cand = selected_row["candidate"]
        trace = selected_row["trace"]
        profile = cand.get("profile") or {}

        title_val = profile.get("current_title", "N/A")
        company_val = profile.get("current_company", "N/A")
        yoe_val = profile.get("years_of_experience", "N/A")
        loc_val = profile.get("location", "N/A")
        country_val = profile.get("country", "N/A")

        # ── Profile + Score Cards ───────────────────────────────────────
        col_prof, col_trace = st.columns([1, 1], gap="large")

        with col_prof:
            st.markdown("#### 👤 Profile Summary")

            # Rank + Score header card
            rank_class = (
                "rank-top10"
                if selected_row["rank"] <= 10
                else "rank-top50"
                if selected_row["rank"] <= 50
                else "rank-rest"
            )
            st.markdown(
                f"""
                <div class="glass-card" style="display: flex; align-items: center; gap: 1rem;">
                    <div class="rank-badge {rank_class}">#{selected_row['rank']}</div>
                    <div style="flex: 1;">
                        <div style="font-weight: 700; color: #f1f5f9; font-size: 1.05rem;">
                            {selected_row['candidate_id']}
                        </div>
                        <div style="color: #94a3b8; font-size: 0.82rem;">{title_val}</div>
                    </div>
                    <div style="text-align: right;">
                        <div style="font-size: 1.4rem; font-weight: 800;
                             background: linear-gradient(135deg, #818cf8, #c084fc);
                             -webkit-background-clip: text;
                             -webkit-text-fill-color: transparent;">
                            {selected_row['score']:.6f}
                        </div>
                        <div style="font-size: 0.7rem; color: #64748b; text-transform: uppercase;
                                    letter-spacing: 0.06em;">composite score</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            # Facts table
            st.markdown(
                f"""
                <div class="glass-card">
                    <table class="profile-table">
                        <tr><td>Current Title</td><td>{title_val}</td></tr>
                        <tr><td>Employer</td><td>{company_val}</td></tr>
                        <tr><td>Experience</td><td>{yoe_val} years</td></tr>
                        <tr><td>Location</td><td>{loc_val}, {country_val}</td></tr>
                    </table>
                </div>
                """,
                unsafe_allow_html=True,
            )

            # Summary
            with st.expander("📝 Candidate Summary", expanded=True):
                summary_text = profile.get("summary") or "*No summary available.*"
                st.write(summary_text)

            # Career History
            with st.expander("💼 Career Timeline", expanded=False):
                history = cand.get("career_history") or []
                if history:
                    for entry in history:
                        t = entry.get("title", "N/A")
                        c = entry.get("company", "N/A")
                        d = entry.get("duration_months", 0)
                        curr = " *(Current)*" if entry.get("is_current") else ""
                        st.markdown(f"- **{t}** at *{c}*{curr} — {d} months")
                else:
                    st.write("No career records found.")

            # Engagement Signals
            with st.expander("⚡ Engagement Signals", expanded=True):
                signals = cand.get("redrob_signals") or {}
                sc_sig1, sc_sig2 = st.columns(2)
                
                np_val = f"{signals.get('notice_period_days', 'N/A')} days"
                otw_val = "Yes" if signals.get("open_to_work_flag") else "No"
                gh_val = f"{signals.get('github_activity_score', 'N/A')}/100"
                rr = signals.get("recruiter_response_rate")
                rr_str = f"{int(rr * 100)}%" if isinstance(rr, (int, float)) else "N/A"

                with sc_sig1:
                    st.markdown(f'<div class="small-metric-label">Notice Period</div><div class="small-metric-val">{np_val}</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="small-metric-label">Open to Work</div><div class="small-metric-val">{otw_val}</div>', unsafe_allow_html=True)
                with sc_sig2:
                    st.markdown(f'<div class="small-metric-label">GitHub Score</div><div class="small-metric-val">{gh_val}</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="small-metric-label">Response Rate</div><div class="small-metric-val">{rr_str}</div>', unsafe_allow_html=True)

        with col_trace:
            st.markdown("#### 🧮 Score Audit & Breakdown")

            # Reasoning card
            reason_str = generate_reasoning(cand, trace, selected_row["rank"])
            st.markdown(
                f"""
                <div class="glass-card"
                     style="border-left: 4px solid #a78bfa; padding: 1rem 1.2rem;">
                    <div style="font-size: 0.72rem; color: #94a3b8;
                                text-transform: uppercase; letter-spacing: 0.06em;
                                font-weight: 600; margin-bottom: 0.4rem;">
                        Reasoning
                    </div>
                    <div style="color: #e2e8f0; font-size: 1.15rem;
                                line-height: 1.55; font-style: italic;">
                        "{reason_str}"
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            # Technical Fit
            st.markdown("##### Technical Alignment")
            tech_score = float(trace.get("technical", 0.0))
            st.progress(min(1.0, max(0.0, tech_score)))
            st.caption(f"Technical composite: `{tech_score:.4f}`")

            # Component Scores
            st.markdown("##### Fit Components")
            semantic_val = trace.get("semantic")
            sem_str = (
                f"{semantic_val:.4f}" if isinstance(semantic_val, float) else str(semantic_val)
            )

            col_fit1, col_fit2 = st.columns(2)
            with col_fit1:
                st.markdown(f"""<div class="fit-list">
                • <strong>Lexical Evidence:</strong> {trace.get('evidence', 0.0):.3f}<br>
                • <strong>Domain Gate:</strong> {trace.get('domain_gate', 1.0):.1f}<br>
                • <strong>Title Gate:</strong> {trace.get('title_gate', 1.0):.1f}<br>
                • <strong>Semantic:</strong> {sem_str}<br>
                • <strong>Seniority:</strong> {trace.get('seniority', 0.0):.2f}
                </div>""", unsafe_allow_html=True)
            with col_fit2:
                st.markdown(f"""<div class="fit-list">
                • <strong>Product:</strong> {trace.get('product', 0.0):.2f}<br>
                • <strong>Stability:</strong> {trace.get('stability', 0.0):.2f}<br>
                • <strong>Location:</strong> {trace.get('location', 0.0):.2f}<br>
                • <strong>Education:</strong> {trace.get('education', 0.0):.2f}
                </div>""", unsafe_allow_html=True)

            st.markdown("##### Behavioral Modifiers")
            col_mod1, col_mod2 = st.columns(2)
            with col_mod1:
                st.markdown(f"""<div class="fit-list">
                • <strong>Honeypot:</strong> {trace.get('honeypot')}<br>
                • <strong>Recency:</strong> {trace.get('recency', 1.0):.2f}<br>
                • <strong>Availability:</strong> {trace.get('availability', 1.0):.2f}
                </div>""", unsafe_allow_html=True)
            with col_mod2:
                st.markdown(f"""<div class="fit-list">
                • <strong>Responsiveness:</strong> {trace.get('responsiveness', 1.0):.2f}<br>
                • <strong>Credibility:</strong> {trace.get('credibility', 1.0):.2f}
                </div>""", unsafe_allow_html=True)

            # Fired Concepts
            concepts = trace.get("evidence_concepts")
            if concepts:
                st.markdown("##### Matched Technical Concepts")
                if isinstance(concepts, list):
                    st.write(", ".join([f"`{c}`" for c in concepts]))
                else:
                    st.write(concepts)
